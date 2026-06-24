from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import csv
from .models import Booking, Resource
from django.contrib.auth.models import User
from django.utils import timezone

class ExportService:
    """Service for exporting data in various formats"""
    
    @staticmethod
    def export_to_excel(report_type='bookings'):
        """Export data to Excel format"""
        wb = Workbook()
        
        if report_type == 'bookings':
            ws = wb.active
            ws.title = 'Bookings'
            
            # Headers
            headers = ['ID', 'Resource', 'Customer', 'Start Time', 'End Time', 'Status', 'Created At']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Data
            for row, booking in enumerate(Booking.objects.select_related('resource', 'customer'), 2):
                ws.cell(row=row, column=1, value=booking.id)
                ws.cell(row=row, column=2, value=booking.resource.name)
                ws.cell(row=row, column=3, value=booking.customer.username)
                ws.cell(row=row, column=4, value=booking.start_time.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=5, value=booking.end_time.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=6, value=booking.get_status_display())
                ws.cell(row=row, column=7, value=booking.created_at.strftime('%Y-%m-%d %H:%M'))
            
            # Adjust column widths
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 20
        
        elif report_type == 'users':
            ws = wb.active
            ws.title = 'Users'
            
            # Headers
            headers = ['Username', 'Email', 'First Name', 'Last Name', 'Date Joined', 'Last Login', 'Is Active']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Data
            for row, user in enumerate(User.objects.all(), 2):
                ws.cell(row=row, column=1, value=user.username)
                ws.cell(row=row, column=2, value=user.email)
                ws.cell(row=row, column=3, value=user.first_name)
                ws.cell(row=row, column=4, value=user.last_name)
                ws.cell(row=row, column=5, value=user.date_joined.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=6, value=user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never')
                ws.cell(row=row, column=7, value='Yes' if user.is_active else 'No')
            
            # Adjust column widths
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 20
        
        elif report_type == 'resources':
            ws = wb.active
            ws.title = 'Resources'
            
            # Headers
            headers = ['ID', 'Name', 'Description', 'Owner', 'Status', 'Category', 'Bookings']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Data
            for row, resource in enumerate(Resource.objects.select_related('owner', 'category'), 2):
                booking_count = Booking.objects.filter(resource=resource).count()
                ws.cell(row=row, column=1, value=resource.id)
                ws.cell(row=row, column=2, value=resource.name)
                ws.cell(row=row, column=3, value=resource.description[:100] + '...' if len(resource.description) > 100 else resource.description)
                ws.cell(row=row, column=4, value=resource.owner.username if resource.owner else 'Unknown')
                ws.cell(row=row, column=5, value=resource.get_status_display())
                ws.cell(row=row, column=6, value=resource.category.name if resource.category else 'Uncategorized')
                ws.cell(row=row, column=7, value=booking_count)
            
            # Adjust column widths
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    @staticmethod
    def export_to_pdf(report_type='bookings'):
        """Export data to PDF format"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a2e'),
            spaceAfter=30
        )
        elements.append(Paragraph(f'{report_type.title()} Report', title_style))
        elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Get data
        if report_type == 'bookings':
            data = [['ID', 'Resource', 'Customer', 'Start Time', 'End Time', 'Status']]
            for booking in Booking.objects.select_related('resource', 'customer')[:50]:
                data.append([
                    str(booking.id),
                    booking.resource.name,
                    booking.customer.username,
                    booking.start_time.strftime('%Y-%m-%d %H:%M'),
                    booking.end_time.strftime('%Y-%m-%d %H:%M'),
                    booking.get_status_display()
                ])
        elif report_type == 'users':
            data = [['Username', 'Email', 'First Name', 'Last Name', 'Date Joined']]
            for user in User.objects.all()[:50]:
                data.append([
                    user.username,
                    user.email,
                    user.first_name,
                    user.last_name,
                    user.date_joined.strftime('%Y-%m-%d %H:%M')
                ])
        else:  # resources
            data = [['Name', 'Description', 'Owner', 'Status', 'Category']]
            for resource in Resource.objects.select_related('owner', 'category')[:50]:
                data.append([
                    resource.name,
                    resource.description[:50] + '...' if len(resource.description) > 50 else resource.description,
                    resource.owner.username if resource.owner else 'Unknown',
                    resource.get_status_display(),
                    resource.category.name if resource.category else 'Uncategorized'
                ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer