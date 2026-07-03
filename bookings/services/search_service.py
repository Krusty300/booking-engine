import logging
import csv
import json
from io import StringIO, BytesIO
from django.db.models import Q, Count
from django.core.cache import cache
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from ..models import Equipment, EquipmentCategory, EquipmentRental, EquipmentReservation, SavedSearch

logger = logging.getLogger(__name__)

class SearchService:
    """Service for advanced equipment search and filtering"""

    @staticmethod
    def search_equipment(request, use_cache=False):
        """
        Perform advanced equipment search with multiple filters
        """
        try:
            # Start with optimized queryset
            queryset = Equipment.objects.select_related(
                'category', 'owner'
            ).prefetch_related(
                'rentals', 'maintenance_records'
            )
            
            # Apply permission-based filtering
            if not request.user.is_staff:
                queryset = queryset.filter(status__in=['AVAILABLE', 'RENTED'])
            
            # Get search parameters
            search_query = request.GET.get('search', '')
            category = request.GET.get('category', '')
            status = request.GET.get('status', '')
            condition = request.GET.get('condition', '')
            location = request.GET.get('location', '')
            min_price = request.GET.get('min_price', '')
            max_price = request.GET.get('max_price', '')
            start_date = request.GET.get('start_date', '')
            end_date = request.GET.get('end_date', '')
            sort_by = request.GET.get('sort', 'name')
            per_page = int(request.GET.get('per_page', 12))
            
            # Limit per_page
            if per_page > 100:
                per_page = 100
            
            # Full-text search
            if search_query:
                queryset = queryset.filter(
                    Q(name__icontains=search_query) |
                    Q(description__icontains=search_query) |
                    Q(serial_number__icontains=search_query) |
                    Q(asset_tag__icontains=search_query) |
                    Q(location__icontains=search_query)
                )
            
            # Category filter
            if category:
                queryset = queryset.filter(category_id=category)
            
            # Status filter
            if status:
                queryset = queryset.filter(status=status)
            
            # Condition filter
            if condition:
                queryset = queryset.filter(condition=condition)
            
            # Location filter
            if location:
                queryset = queryset.filter(location__icontains=location)
            
            # Price range filter
            if min_price:
                try:
                    queryset = queryset.filter(purchase_price__gte=float(min_price))
                except ValueError:
                    pass
            if max_price:
                try:
                    queryset = queryset.filter(purchase_price__lte=float(max_price))
                except ValueError:
                    pass
            
            # Availability date range filter
            if start_date and end_date:
                try:
                    start_date = timezone.datetime.fromisoformat(start_date)
                    end_date = timezone.datetime.fromisoformat(end_date)
                    
                    if timezone.is_naive(start_date):
                        start_date = timezone.make_aware(start_date)
                    if timezone.is_naive(end_date):
                        end_date = timezone.make_aware(end_date)
                    
                    # Check for rentals
                    rented_ids = EquipmentRental.objects.filter(
                        status='CHECKED_OUT',
                        checkout_date__lt=end_date,
                        expected_return_date__gt=start_date
                    ).values_list('equipment_id', flat=True)
                    
                    # Check for reservations
                    reserved_ids = EquipmentReservation.objects.filter(
                        status__in=['PENDING', 'CONFIRMED'],
                        start_date__lt=end_date,
                        end_date__gt=start_date
                    ).values_list('equipment_id', flat=True)
                    
                    # Combine unavailable IDs
                    unavailable_ids = set(rented_ids) | set(reserved_ids)
                    queryset = queryset.exclude(id__in=unavailable_ids)
                except (ValueError, TypeError):
                    pass
            
            # Sorting (fixed)
            sort_options = {
                'name': 'name',
                '-name': '-name',
                'created_at': 'created_at',
                '-created_at': '-created_at',
                'price': 'purchase_price',
                '-price': '-purchase_price',
                'status': 'status',
                '-status': '-status',
                'condition': 'condition',
                '-condition': '-condition',
            }
            
            if sort_by in sort_options:
                queryset = queryset.order_by(sort_options[sort_by])
            else:
                queryset = queryset.order_by('name')
            
            # Get total count before pagination
            total_count = queryset.count()
            
            # Pagination
            paginator = Paginator(queryset, per_page)
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
            
            # Track search analytics
            if search_query and len(search_query) >= 2:
                try:
                    from ..models import AnalyticsEvent
                    AnalyticsEvent.objects.create(
                        user=request.user if request.user.is_authenticated else None,
                        event_type='SEARCH',
                        url=request.path,
                        metadata={
                            'query': search_query,
                            'results': total_count,
                            'filters': SearchService._get_active_filters(request)
                        }
                    )
                except Exception as e:
                    logger.warning(f"Failed to log search analytics: {e}")
            
            # Get search metadata
            search_metadata = {
                'total_results': total_count,
                'search_query': search_query,
                'filters_applied': SearchService._get_active_filters(request),
                'per_page': per_page,
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
            }
            
            logger.info(f"Search performed: '{search_query}' - {total_count} results")
            
            return {
                'results': page_obj,
                'metadata': search_metadata,
                'queryset': queryset,
            }
            
        except Exception as e:
            logger.error(f"Search error: {e}", exc_info=True)
            raise
    
    @staticmethod
    def get_filter_options(use_cache=True):
        """Get all filter options for the search form with caching"""
        cache_key = 'filter_options'
        
        if use_cache:
            options = cache.get(cache_key)
            if options:
                logger.info("Filter options served from cache")
                return options
        
        try:
            options = {
                'categories': EquipmentCategory.objects.all().order_by('name'),
                'statuses': Equipment.STATUS_CHOICES,
                'conditions': Equipment.CONDITION_CHOICES,
                'locations': Equipment.objects.values_list(
                    'location', flat=True
                ).distinct().exclude(location='').order_by('location'),
            }
            
            cache.set(cache_key, options, 3600)  # Cache for 1 hour
            logger.info("Filter options generated and cached")
            return options
            
        except Exception as e:
            logger.error(f"Error getting filter options: {e}", exc_info=True)
            raise
    
    @staticmethod
    def get_facets(queryset):
        """Get facet counts for search results"""
        try:
            return {
                'categories': queryset.values('category__name').annotate(
                    count=Count('category')
                ).order_by('-count')[:10],
                'statuses': queryset.values('status').annotate(
                    count=Count('status')
                ).order_by('-count'),
                'conditions': queryset.values('condition').annotate(
                    count=Count('condition')
                ).order_by('-count'),
                'locations': queryset.values('location').annotate(
                    count=Count('location')
                ).exclude(location='').order_by('-count')[:10],
            }
        except Exception as e:
            logger.error(f"Error getting facets: {e}", exc_info=True)
            return {}
    
    @staticmethod
    def get_spell_correction(query):
        """Get spell correction suggestions"""
        from difflib import get_close_matches
        
        if not query or len(query) < 3:
            return None
        
        try:
            # Get common equipment names
            common_names = Equipment.objects.values_list('name', flat=True).distinct()
            matches = get_close_matches(query, common_names, n=1, cutoff=0.8)
            return matches[0] if matches else None
        except Exception as e:
            logger.error(f"Error getting spell correction: {e}")
            return None
    
    @staticmethod
    def _get_active_filters(request):
        """Get dictionary of active filters"""
        filters = {}
        filter_fields = ['category', 'status', 'condition', 'location', 'min_price', 'max_price', 'start_date', 'end_date']
        
        for field in filter_fields:
            value = request.GET.get(field, '')
            if value:
                filters[field] = value
        
        return filters
    
    @staticmethod
    def get_search_suggestions(query):
        """Get autocomplete suggestions for search"""
        if not query or len(query) < 2:
            return []
        
        suggestions = []
        
        try:
            # Equipment name suggestions
            name_results = Equipment.objects.filter(
                name__icontains=query
            ).values_list('name', flat=True).distinct()[:5]
            for result in name_results:
                suggestions.append({
                    'type': 'equipment',
                    'label': result,
                    'value': result,
                    'category': 'Equipment'
                })
            
            # Serial number suggestions
            serial_results = Equipment.objects.filter(
                serial_number__icontains=query
            ).values_list('serial_number', flat=True).distinct()[:3]
            for result in serial_results:
                suggestions.append({
                    'type': 'serial',
                    'label': f'Serial: {result}',
                    'value': result,
                    'category': 'Serial Number'
                })
            
            # Category suggestions
            category_results = EquipmentCategory.objects.filter(
                name__icontains=query
            )[:3]
            for result in category_results:
                suggestions.append({
                    'type': 'category',
                    'label': f'Category: {result.name}',
                    'value': result.name,
                    'category': 'Category',
                    'id': result.id
                })
            
            # Location suggestions
            location_results = Equipment.objects.filter(
                location__icontains=query
            ).values_list('location', flat=True).distinct()[:3]
            for result in location_results:
                suggestions.append({
                    'type': 'location',
                    'label': f'Location: {result}',
                    'value': result,
                    'category': 'Location'
                })
            
        except Exception as e:
            logger.error(f"Error getting search suggestions: {e}")
        
        return suggestions[:12]

    @staticmethod
    def save_search(user, name, request):
        """Save a search"""
        try:
            # Get all search parameters
            filters = {}
            filter_fields = ['category', 'status', 'condition', 'location', 'min_price', 'max_price', 'start_date', 'end_date']
            
            for field in filter_fields:
                value = request.GET.get(field, '')
                if value:
                    filters[field] = value
            
            # Get sorting and pagination
            sort_by = request.GET.get('sort', 'name')
            per_page = int(request.GET.get('per_page', 12))
            search_query = request.GET.get('search', '')
            
            # Check if search with this name already exists
            saved_search, created = SavedSearch.objects.get_or_create(
                user=user,
                name=name,
                defaults={
                    'search_query': search_query,
                    'filters': filters,
                    'sort_by': sort_by,
                    'per_page': per_page,
                }
            )
            
            if not created:
                # Update existing search
                saved_search.search_query = search_query
                saved_search.filters = filters
                saved_search.sort_by = sort_by
                saved_search.per_page = per_page
                saved_search.save()
            
            logger.info(f"Search saved: {name} for user {user.username}")
            return saved_search, created
            
        except Exception as e:
            logger.error(f"Error saving search: {e}", exc_info=True)
            raise

    @staticmethod
    def get_saved_searches(user):
        """Get all saved searches for a user"""
        return SavedSearch.objects.filter(user=user).order_by('-created_at')

    @staticmethod
    def delete_saved_search(user, search_id):
        """Delete a saved search"""
        search = get_object_or_404(SavedSearch, id=search_id, user=user)
        search.delete()
        logger.info(f"Search deleted: {search_id} for user {user.username}")
        return True

    @staticmethod
    def toggle_favorite(user, search_id):
        """Toggle favorite status of a saved search"""
        search = get_object_or_404(SavedSearch, id=search_id, user=user)
        search.is_favorite = not search.is_favorite
        search.save()
        logger.info(f"Search favorite toggled: {search_id} - {search.is_favorite}")
        return search.is_favorite

    @staticmethod
    def export_search_results(request):
        """Export search results in various formats"""
        try:
            format_type = request.GET.get('format', 'csv')
            
            # Get search results
            search_data = SearchService.search_equipment(request)
            queryset = search_data['queryset']
            
            if format_type == 'csv':
                return SearchService._export_csv(queryset)
            elif format_type == 'json':
                return SearchService._export_json(queryset)
            elif format_type == 'excel':
                return SearchService._export_excel(queryset)
            elif format_type == 'pdf':
                return SearchService._export_pdf(queryset)
            else:
                return None
                
        except Exception as e:
            logger.error(f"Error exporting search results: {e}", exc_info=True)
            raise

    @staticmethod
    def _export_csv(queryset):
        """Export to CSV format"""
        try:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = (
                f'attachment; filename="equipment_search_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            )
            
            writer = csv.writer(response)
            writer.writerow([
                'ID', 'Name', 'Serial Number', 'Category', 'Status', 
                'Condition', 'Location', 'Purchase Price', 'Purchase Date',
                'Warranty Expiry', 'Created At'
            ])
            
            for equipment in queryset.iterator(chunk_size=1000):
                writer.writerow([
                    equipment.id,
                    equipment.name,
                    equipment.serial_number,
                    equipment.category.name if equipment.category else '',
                    equipment.get_status_display(),
                    equipment.get_condition_display(),
                    equipment.location or '',
                    equipment.purchase_price or '',
                    equipment.purchase_date.strftime('%Y-%m-%d') if equipment.purchase_date else '',
                    equipment.warranty_expiry.strftime('%Y-%m-%d') if equipment.warranty_expiry else '',
                    equipment.created_at.strftime('%Y-%m-%d %H:%M'),
                ])
            
            logger.info(f"CSV exported with {queryset.count()} rows")
            return response
            
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}", exc_info=True)
            raise

    @staticmethod
    def _export_json(queryset):
        """Export to JSON format"""
        try:
            data = []
            for equipment in queryset:
                data.append({
                    'id': equipment.id,
                    'name': equipment.name,
                    'serial_number': equipment.serial_number,
                    'category': equipment.category.name if equipment.category else None,
                    'status': equipment.status,
                    'condition': equipment.condition,
                    'location': equipment.location,
                    'purchase_price': float(equipment.purchase_price) if equipment.purchase_price else None,
                    'purchase_date': equipment.purchase_date.isoformat() if equipment.purchase_date else None,
                    'warranty_expiry': equipment.warranty_expiry.isoformat() if equipment.warranty_expiry else None,
                    'created_at': equipment.created_at.isoformat(),
                })
            
            response = JsonResponse({
                'results': data,
                'total': queryset.count(),
                'exported_at': timezone.now().isoformat(),
            })
            
            filename = f"equipment_search_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            logger.info(f"JSON exported with {queryset.count()} rows")
            return response
            
        except Exception as e:
            logger.error(f"Error exporting JSON: {e}", exc_info=True)
            raise

    @staticmethod
    def _export_excel(queryset):
        """Export to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return SearchService._export_csv(queryset)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Equipment Search Results"
            
            # Headers
            headers = ['ID', 'Name', 'Serial Number', 'Category', 'Status', 
                       'Condition', 'Location', 'Purchase Price', 'Purchase Date',
                       'Warranty Expiry', 'Created At']
            
            # Style headers
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            for row, equipment in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=equipment.id)
                ws.cell(row=row, column=2, value=equipment.name)
                ws.cell(row=row, column=3, value=equipment.serial_number)
                ws.cell(row=row, column=4, value=equipment.category.name if equipment.category else '')
                ws.cell(row=row, column=5, value=equipment.get_status_display())
                ws.cell(row=row, column=6, value=equipment.get_condition_display())
                ws.cell(row=row, column=7, value=equipment.location or '')
                ws.cell(row=row, column=8, value=float(equipment.purchase_price) if equipment.purchase_price else '')
                ws.cell(row=row, column=9, value=equipment.purchase_date.strftime('%Y-%m-%d') if equipment.purchase_date else '')
                ws.cell(row=row, column=10, value=equipment.warranty_expiry.strftime('%Y-%m-%d') if equipment.warranty_expiry else '')
                ws.cell(row=row, column=11, value=equipment.created_at.strftime('%Y-%m-%d %H:%M'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to response
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="equipment_search_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            )
            
            logger.info(f"Excel exported with {queryset.count()} rows")
            return response
            
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}", exc_info=True)
            raise

    @staticmethod
    def _export_pdf(queryset):
        """Export to PDF format"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
        except ImportError:
            logger.warning("reportlab not installed, falling back to CSV")
            return SearchService._export_csv(queryset)
        
        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = (
                f'attachment; filename="equipment_search_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            )
            
            # Create PDF document
            doc = SimpleDocTemplate(response, pagesize=landscape(A4))
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            
            # Title
            elements.append(Paragraph("Equipment Search Results", title_style))
            elements.append(Paragraph(f"Generated on {timezone.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Table data (limit to 50 items for PDF)
            data = [['ID', 'Name', 'Serial', 'Category', 'Status', 'Condition', 'Location', 'Price']]
            
            for equipment in queryset[:50]:
                data.append([
                    str(equipment.id),
                    equipment.name[:30],
                    equipment.serial_number,
                    equipment.category.name if equipment.category else '',
                    equipment.get_status_display(),
                    equipment.get_condition_display(),
                    equipment.location[:20] if equipment.location else '',
                    f"${equipment.purchase_price}" if equipment.purchase_price else ''
                ])
            
            # Create table
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            logger.info(f"PDF exported with {min(queryset.count(), 50)} rows")
            return response
            
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}", exc_info=True)
            raise